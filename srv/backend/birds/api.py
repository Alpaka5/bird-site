import logging
from urllib.request import Request

from ninja import Router, ModelSchema, Schema, File
from ninja.files import UploadedFile
from import_export import resources
import tablib
import openpyxl

import birds.models as models
from birds import resources as bird_resources


router = Router()


class BirdSchema(ModelSchema):
    class Meta:
        model = models.Bird
        fields = "__all__"


class BirdDescriptionSchema(ModelSchema):
    class Meta:
        model = models.BirdTextField
        fields = "__all__"


class Error(Schema):
    message: str


@router.get("/by_latin_name/{bird_latin_name}", response={200: BirdSchema, 404: Error})
def get_bird_by_id(
    request,
    bird_latin_name: str,
):
    try:
        return 200, models.Bird.objects.get(latin_name=bird_latin_name)
    except models.Bird.DoesNotExist:
        return 404, {"message": f"Bird with name: {bird_latin_name} doesn't exist"}


@router.get("/all_birds", response={200: list[BirdSchema], 404: Error})
def get_all_birds(request):
    logging.info("DOWNLOADING BIRDS DATA")
    return (200, models.Bird.objects.all())


@router.get(
    "/description/{bird_latin_name}/{language_id}",
    response={200: BirdDescriptionSchema, 404: Error},
)
def get_bird_description(request, bird_latin_name: str, language_id: str):
    logging.info(f"Fetching {bird_latin_name} description")
    languages = list(models.Language.objects.all().values_list(flat=True))
    if language_id not in languages:
        return (
            404,
            {"message": f"Language {language_id} not found, use one of {languages}"},
        )
    try:
        fetched_bird = models.BirdTextField.objects.get(
            bird_id=bird_latin_name, language_id=language_id
        )
    except models.BirdTextField.DoesNotExist as err:
        return (
            200,
            {
                "bird_id": bird_latin_name,
                "language_id": language_id,
                "description": f"Temporary description generated for {bird_latin_name} because I was lazy for now",
            },
        )
    logging.info(fetched_bird)
    return (200, fetched_bird)


class UpdateResponse(Schema):
    message: str


@router.post("/db/update", response={200: UpdateResponse, 400: Error, 422: Error})
def update_db_with_birds(request: Request, file: UploadedFile = File(...)):
    if not file.name.endswith(".xlsx"):
        return 422, Error(message="Unsupported file format. Please upload a .xlsx file")

    workbook: openpyxl.Workbook = openpyxl.load_workbook(file)
    if not all(
        [
            (sheet.lower() in models.mapped_models.keys())
            for sheet in workbook.sheetnames
        ]
    ):
        return 400, Error(message="Excel file contains invalid sheet names")

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        db_fields = list(
            list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        )

        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            model_resource = resources.modelresource_factory(
                model=models.mapped_models[sheet_name.lower()],
                resource_class=bird_resources.BaseResourceClass,
            )
            model_resource = model_resource()
            model_resource.import_data(
                dataset=tablib.Dataset(list(row_values), headers=db_fields),
                raise_errors=True,
            )

    return 200, UpdateResponse(message="Excel file was uploaded successfully!")
