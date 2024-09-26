import openpyxl
from fastapi import APIRouter, UploadFile, HTTPException

router = APIRouter()


@router.post("/db/update", tags=["db"])
def update_db_with_birds(file: UploadFile):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=422,
            detail="Unsupported file format. Please upload a .xlsx file",
        )

    workbook: openpyxl.Workbook = openpyxl.load_workbook(file.file)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        db_fields = list(
            list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        )

        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            model_resource = resources.modelresource_factory(
                model=models.mapped_models[sheet_name.lower()],
                resource_class=bird_resources.BaseResourceClass,
            )
            model_resource = model_resource()
            model_resource.import_data(
                dataset=tablib.Dataset(list(row_values), headers=db_fields),
                raise_errors=True,
            )

    return 200, UpdateResponse(message="Excel file was uploaded successfully!")
