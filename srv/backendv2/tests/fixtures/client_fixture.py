import openpyxl
from pytest import fixture

from fastapi.testclient import TestClient
from sqlalchemy import create_engine, StaticPool
from sqlalchemy.orm import sessionmaker

from database import models
from database.database import Base, get_db
from main import app


@fixture(scope="session")
def db_session():
    SQLALCHEMY_DATABASE_URL = "sqlite://"

    engine = create_engine(
        SQLALCHEMY_DATABASE_URL,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

    Base.metadata.create_all(bind=engine)

    return TestingSessionLocal()


@fixture(scope="session")
def loaded_db(db_session):

    with open("tests/fixtures/test_birds_data.xlsx", "rb") as test_bird_data:

        workbook: openpyxl.Workbook = openpyxl.load_workbook(test_bird_data)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        current_model = getattr(models, sheet_name)
        db_fields = list(
            list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        )

        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = {field: value for field, value in zip(db_fields, row_values)}
            db_session.add(current_model(**row_dict))

    db_session.commit()

    return db_session


@fixture(scope="session")
def loaded_api_test_client(loaded_db):
    def override_get_db():
        try:
            yield loaded_db
        finally:
            loaded_db.close()

    app.dependency_overrides[get_db] = override_get_db

    return TestClient(app)


@fixture(scope="session")
def api_test_client():
    SQLALCHEMY_DATABASE_URL = "sqlite://"

    engine = create_engine(
        SQLALCHEMY_DATABASE_URL,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

    Base.metadata.create_all(bind=engine)

    def override_get_db():
        try:
            db = TestingSessionLocal()
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db

    return TestClient(app)
