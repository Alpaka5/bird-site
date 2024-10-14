import openpyxl
from fastapi import APIRouter, UploadFile, HTTPException, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from database.database import get_db
from database import schemas, models, crud
from helper.logger import logger

router = APIRouter(prefix="/birds")


@router.get("/by_latin_name/{bird_latin_name}", response_model=schemas.Bird)
def get_bird_by_id(bird_latin_name: str, db: Session = Depends(get_db)):

    bird = crud.get_single_bird(db, bird_latin_name)
    if bird:
        return bird[0]
    else:
        raise HTTPException(
            status_code=404, detail=f"Bird with name: {bird_latin_name} doesn't exist"
        )


@router.get("/all_birds", response_model=list[schemas.Bird])
def get_all_birds(db: Session = Depends(get_db)):
    logger.info("DOWNLOADING BIRDS DATA")
    return crud.get_all_birds(db)


@router.get(
    "/description/{bird_latin_name}/{language_id}", response_model=schemas.BirdTextField
)
def get_bird_description(
    bird_latin_name: str, language_id: str, db: Session = Depends(get_db)
):
    logger.info(f"Fetching {bird_latin_name} description")
    languages = [language.code for language in crud.get_supported_languages(db)]
    if language_id not in languages:
        raise HTTPException(
            status_code=404,
            detail=f"Language {language_id} not found, use one of {languages}",
        )

    fetched_bird = crud.get_bird_text_field(
        db=db, bird_latin_name=bird_latin_name, language_id=language_id
    )

    if fetched_bird is None:
        return {
            "bird": bird_latin_name,
            "language": language_id,
            "description": f"Temporary description generated for {bird_latin_name} because I was lazy for now",
        }

    logger.info(fetched_bird[0])
    return fetched_bird[0]


@router.post("/db/update", tags=["db"])
def update_db_with_birds(file: UploadFile, db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=422,
            detail="Unsupported file format. Please upload a .xlsx file",
        )

    workbook: openpyxl.Workbook = openpyxl.load_workbook(file.file)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        current_model = getattr(models, sheet_name)
        db_fields = list(
            list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        )

        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = {field: value for field, value in zip(db_fields, row_values)}
            db.add(current_model(**row_dict))

    db.commit()

    return {"message": "WORKED HELLA FINE"}


@router.get("/image/{bird_latin_name}")
def get_bird_image(bird_latin_name):
    return FileResponse(
        f"assets/images/{bird_latin_name.replace(' ','_')}.png", media_type="image/png"
    )


@router.get("/sound/{bird_latin_name}")
def get_bird_image(bird_latin_name):
    return FileResponse(
        f"assets/sounds/{bird_latin_name.replace(' ','_')}.m4a", media_type="audio/m4a"
    )
